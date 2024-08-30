import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Read the text file
df = pd.read_csv('newgg.txt', delimiter = ', ', header=None, engine='python')
# Define column names
df.columns = ["ID", "FIRST", "AMOUNT", "WEEK", "LAST"]

# Remove column names from cell values
for col in df.columns:
    df[col] = df[col].str.split(":").str[1].str.strip()

# Add new column "GROUP" with value "L831" for all rows
df['GROUP'] = 'COBABW'


# Reorder the columns
df = df.reindex(columns=["ID", "LAST", "FIRST", "AMOUNT", "WEEK", "GROUP"])

# Convert ID to string format to preserve leading zeros
df['ID'] = df['ID'].astype(str)

# Convert 'AMOUNT' to numeric format
df['AMOUNT'] = pd.to_numeric(df['AMOUNT'], errors='coerce')

# Define a function to format a number as a currency
def format_currency(num):
    return "${:,.2f}".format(num)

# Overwrite the 'AMOUNT' column with the formatted values
df['AMOUNT'] = df['AMOUNT'].apply(format_currency)

# Reorder the columns
df = df.reindex(columns=["ID", "LAST", "FIRST", "AMOUNT", "WEEK", "GROUP"])

# Write to an Excel file
df.to_excel('COBA BW 08-30-24 PAYFILE_EXTRACTED.xlsx', index=False)
# Load the workbook and select the first sheet
wb = load_workbook('COBA BW 08-30-24 PAYFILE_EXTRACTED.xlsx')
sheet = wb.active

# Adjust the width of the columns
sheet.column_dimensions[get_column_letter(2)].width = 20  # LAST column
sheet.column_dimensions[get_column_letter(3)].width = 20  # FIRST column
sheet.column_dimensions[get_column_letter(5)].width = 20  # WEEK column

# Save the changes
wb.save('COBA BW 08-30-24 PAYFILE_EXTRACTED.xlsx')